"""
route_optimizer.py
------------------
End-to-end delivery route optimizer using:
  - US Census Geocoder (free, no API key) for geocoding
  - Haversine formula for distance matrix (free, local)
  - Nearest-neighbor + 2-opt for route optimization (free, local)

Usage:
    python route_optimizer.py --input ../RuralMapping.xlsx --output route.csv

Input file can be CSV or XLSX and must have columns that map to:
street, city, zip (state optional)
Output CSV contains the original rows in optimized delivery order.

To manually supply coordinates for addresses that fail geocoding, add
Latitude and Longitude columns to the input file. Rows with both values
present will skip the geocoder entirely on the next run.

Optional flags:
    --start-index   Row index (0-based) to use as the starting point (default: 0)
    --no-two-opt    Skip 2-opt improvement pass (faster, lower quality)
    --nominatim     Use Nominatim (OSM) geocoder instead of Census (for non-US addresses)
    --map-output    Output HTML map file name (default: route_map.html)
"""

import argparse
import csv
import json
import io
import math
import os
import sys
import time
from typing import Optional

from openpyxl import load_workbook
import requests


# ---------------------------------------------------------------------------
# 1. Geocoding
# ---------------------------------------------------------------------------

CENSUS_GEOCODE_URL = (
    "https://geocoding.geo.census.gov/geocoder/locations/addressbatch"
)
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"


def geocode_census(addresses: list[dict]) -> list[Optional[tuple[float, float]]]:
    """
    Geocode a list of US addresses using the Census Bureau batch API.
    Returns a list of (lat, lon) tuples, or None for addresses that failed.

    addresses: list of dicts with keys: street, city, state, zip
    """
    print(f"  Geocoding {len(addresses)} addresses via US Census Geocoder...")

    # Build the CSV payload expected by the Census API
    lines = []
    for i, a in enumerate(addresses):
        street = a.get("street", "").replace('"', "'")
        city   = a.get("city",   "").replace('"', "'")
        state  = a.get("state",  "").replace('"', "'")
        zip_   = a.get("zip",    "").replace('"', "'")
        lines.append(f'{i},"{street}","{city}","{state}","{zip_}"')

    payload_csv = "\n".join(lines)

    response = requests.post(
        CENSUS_GEOCODE_URL,
        files={"addressFile": ("addresses.csv", payload_csv, "text/csv")},
        data={
            "benchmark": "Public_AR_Current",
            "vintage": "Current_Current",
        },
        timeout=120,
    )
    response.raise_for_status()

    # The response is a CSV. Column layout for matched rows:
    #   0: input ID
    #   1: input address
    #   2: match status ("Match" / "No_Match" / "Tie")
    #   3: match type ("Exact" / "Non_Exact")
    #   4: matched address string
    #   5: coordinates as a SINGLE "lon,lat" string (e.g. "-77.0366,38.8977")
    #   6: Tiger/Line ID
    #   7: side of street
    results = [None] * len(addresses)
    reader = csv.reader(io.StringIO(response.text))
    for row in reader:
        if len(row) < 6:
            continue
        try:
            idx = int(row[0])
            match_status = row[2].strip()  # "Match" or "No_Match"
            if match_status == "Match" and row[5]:
                coord_parts = row[5].strip().split(",")
                lon = float(coord_parts[0])
                lat = float(coord_parts[1])
                results[idx] = (lat, lon)
        except (ValueError, IndexError):
            continue

    matched = sum(1 for r in results if r is not None)
    print(f"  Geocoded {matched}/{len(addresses)} addresses successfully.")
    if matched < len(addresses):
        unmatched = [i for i, r in enumerate(results) if r is None]
        print(f"  WARNING: {len(unmatched)} addresses could not be geocoded: "
              f"rows {unmatched[:10]}{'...' if len(unmatched) > 10 else ''}")
    return results


def geocode_nominatim(addresses: list[dict]) -> list[Optional[tuple[float, float]]]:
    """
    Geocode addresses using Nominatim (OpenStreetMap). Works for non-US addresses.
    Rate-limited to 1 request/second per OSM usage policy.
    633 addresses will take ~10-11 minutes.
    """
    print(f"  Geocoding {len(addresses)} addresses via Nominatim (OSM)...")
    print("  Note: rate-limited to 1 req/sec — this will take several minutes.")

    results = []
    for i, a in enumerate(addresses):
        query = ", ".join(filter(None, [
            a.get("street", ""),
            a.get("city",   ""),
            a.get("state",  ""),
            a.get("zip",    ""),
            a.get("country", ""),
        ]))
        try:
            resp = requests.get(
                NOMINATIM_URL,
                params={"q": query, "format": "json", "limit": 1},
                headers={"User-Agent": "route-optimizer/1.0"},
                timeout=10,
            )
            resp.raise_for_status()
            data = resp.json()
            if data:
                lat = float(data[0]["lat"])
                lon = float(data[0]["lon"])
                results.append((lat, lon))
            else:
                results.append(None)
        except Exception as e:
            print(f"  Row {i} geocoding error: {e}")
            results.append(None)

        if (i + 1) % 50 == 0:
            print(f"  ...{i + 1}/{len(addresses)} geocoded")

        time.sleep(1.0)  # OSM usage policy: max 1 req/sec

    matched = sum(1 for r in results if r is not None)
    print(f"  Geocoded {matched}/{len(addresses)} addresses successfully.")
    return results


# ---------------------------------------------------------------------------
# 2. Distance Matrix
# ---------------------------------------------------------------------------

def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Straight-line distance in miles between two lat/lon points."""
    R = 3958.8  # Earth radius in miles
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def build_distance_matrix(coords: list[tuple[float, float]]) -> list[list[float]]:
    """Build a full NxN haversine distance matrix. Runs locally in < 1 second."""
    n = len(coords)
    print(f"  Building {n}x{n} haversine distance matrix...")
    matrix = [[0.0] * n for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            d = haversine(coords[i][0], coords[i][1], coords[j][0], coords[j][1])
            matrix[i][j] = d
            matrix[j][i] = d
    print(f"  Distance matrix complete ({n*n:,} entries).")
    return matrix


# ---------------------------------------------------------------------------
# 3. Route Optimization
# ---------------------------------------------------------------------------

def nearest_neighbor(dist_matrix: list[list[float]], start: int = 0) -> list[int]:
    """
    Greedy nearest-neighbor heuristic. O(n²).
    Returns a list of indices in visit order.
    """
    n = len(dist_matrix)
    unvisited = set(range(n))
    route = [start]
    unvisited.remove(start)
    while unvisited:
        last = route[-1]
        nearest = min(unvisited, key=lambda x: dist_matrix[last][x])
        route.append(nearest)
        unvisited.remove(nearest)
    return route


def two_opt(route: list[int], dist_matrix: list[list[float]]) -> list[int]:
    """
    2-opt improvement pass. Repeatedly reverses route segments when doing
    so reduces total distance. Runs until no improvement is found.
    Typically brings routes within 5-15% of optimal for real-world datasets.
    """
    n = len(route)
    improved = True
    best_route = route[:]
    iteration = 0

    while improved:
        improved = False
        iteration += 1
        for i in range(1, n - 1):
            for j in range(i + 1, n):
                # Cost before: ...→route[i-1]→route[i]→...→route[j-1]→route[j]→...
                # Cost after:  ...→route[i-1]→route[j-1]→...→route[i]→route[j]→...
                before = (dist_matrix[best_route[i - 1]][best_route[i]] +
                          dist_matrix[best_route[j - 1]][best_route[j % n]])
                after  = (dist_matrix[best_route[i - 1]][best_route[j - 1]] +
                          dist_matrix[best_route[i]][best_route[j % n]])
                if after < before - 1e-10:
                    best_route[i:j] = best_route[i:j][::-1]
                    improved = True
        print(f"  2-opt iteration {iteration} complete.")

    return best_route


def route_total_distance(route: list[int], dist_matrix: list[list[float]]) -> float:
    return sum(dist_matrix[route[i]][route[i + 1]] for i in range(len(route) - 1))


# ---------------------------------------------------------------------------
# 4. I/O Helpers
# ---------------------------------------------------------------------------

def normalize_column_name(name: str) -> str:
    """Normalize a column name for case-insensitive matching across CSV/XLSX headers."""
    return "".join(ch for ch in name.strip().lower() if ch.isalnum())


def canonicalize_row(row: dict[str, str]) -> dict[str, str]:
    """Map common address header variants to canonical keys expected by geocoders."""
    aliases = {
        "street": {
            "street",
            "address",
            "address1",
            "addressline1",
            "streetaddress",
            "streetaddr",
            "residentialaddress",
            "homeaddress",
        },
        "city": {"city", "town", "municipality"},
        "state": {"state", "province", "region", "st"},
        "zip": {"zip", "zipcode", "postal", "postalcode", "postcode"},
        "country": {"country"},
    }

    canonical = row.copy()
    for target, keys in aliases.items():
        for key in keys:
            if key in row and row[key]:
                canonical[target] = row[key]
                break
        else:
            canonical.setdefault(target, "")

    return canonical


def load_rows_from_csv(path: str) -> tuple[list[dict], list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows = list(reader)

    normalized = []
    for row in rows:
        normalized_row = {}
        for k, v in row.items():
            if k is None:
                continue
            normalized_row[normalize_column_name(k)] = str(v).strip() if v is not None else ""
        normalized.append(canonicalize_row(normalized_row))

    return normalized, [f.strip() for f in fieldnames]


def load_rows_from_xlsx(path: str) -> tuple[list[dict], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)

    headers = next(row_iter, None)
    if headers is None:
        return [], []

    fieldnames = [str(h).strip() if h is not None else "" for h in headers]
    rows: list[dict] = []

    for values in row_iter:
        if values is None:
            continue
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        normalized_row = {}
        for header, value in zip(fieldnames, values):
            if not header:
                continue
            normalized_row[normalize_column_name(header)] = str(value).strip() if value is not None else ""
        rows.append(canonicalize_row(normalized_row))

    return rows, fieldnames


def load_addresses(path: str) -> tuple[list[dict], list[str]]:
    """Load addresses from CSV or XLSX. Returns (rows, fieldnames)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        normalized, fieldnames = load_rows_from_xlsx(path)
    elif ext == ".csv":
        normalized, fieldnames = load_rows_from_csv(path)
    else:
        raise ValueError(f"Unsupported input file type: {ext}. Use .csv or .xlsx")

    required = {"street", "city", "zip"}
    found = set(normalized[0].keys()) if normalized else set()
    missing = required - found
    if missing:
        raise ValueError(
            f"Input file is missing required columns: {missing}\n"
            f"Found columns: {found}\n"
            f"Expected at minimum: street, city, zip (state is optional)"
        )

    return normalized, fieldnames


def extract_manual_coords(addresses: list[dict]) -> list[Optional[tuple[float, float]]]:
    """
    Read pre-supplied coordinates from address rows.
    Expects canonical keys 'latitude' and 'longitude' (populated when the input
    file contains Latitude/Longitude columns).
    Returns None for any row missing either value or where parsing fails.
    """
    result: list[Optional[tuple[float, float]]] = []
    for addr in addresses:
        lat_str = addr.get("latitude", "").strip()
        lon_str = addr.get("longitude", "").strip()
        if lat_str and lon_str:
            try:
                result.append((float(lat_str), float(lon_str)))
            except ValueError:
                result.append(None)
        else:
            result.append(None)
    return result


def save_route(
    path: str,
    addresses: list[dict],
    route: list[int],
    coords: list[Optional[tuple[float, float]]],
    fieldnames: list[str],
):
    """Write optimized route to CSV, preserving original columns + adding lat/lon/stop_number."""
    out_fieldnames = ["stop_number"] + fieldnames + ["latitude", "longitude"]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=out_fieldnames)
        writer.writeheader()
        for stop_num, idx in enumerate(route, start=1):
            row = {"stop_number": stop_num}
            addr = addresses[idx]
            for field in fieldnames:
                row[field] = addr.get(normalize_column_name(field), "")
            if coords[idx]:
                row["latitude"]  = round(coords[idx][0], 6)
                row["longitude"] = round(coords[idx][1], 6)
            else:
                row["latitude"]  = ""
                row["longitude"] = ""
            writer.writerow(row)

    print(f"  Route saved to: {path}")


def save_leaflet_map(
        path: str,
        addresses: list[dict],
        route: list[int],
        coords: list[Optional[tuple[float, float]]],
):
        """Write an interactive Leaflet map HTML for the optimized route."""
        points = []
        for stop_num, idx in enumerate(route, start=1):
                coord = coords[idx]
                if coord is None:
                        continue

                addr = addresses[idx]
                address_parts = [
                        addr.get("street", ""),
                        addr.get("city", ""),
                        addr.get("state", ""),
                        addr.get("zip", ""),
                ]
                address_text = ", ".join(part for part in address_parts if part)

                points.append({
                        "stop": stop_num,
                        "lat": round(coord[0], 6),
                        "lon": round(coord[1], 6),
                        "address": address_text,
                })

        if not points:
                print("  Map skipped: no geocoded points available.")
                return

        center_lat = sum(p["lat"] for p in points) / len(points)
        center_lon = sum(p["lon"] for p in points) / len(points)
        points_json = json.dumps(points)

        html_doc = f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
    <meta charset=\"utf-8\" />
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
    <title>Optimized Route Map</title>
    <link
        rel=\"stylesheet\"
        href=\"https://unpkg.com/leaflet@1.9.4/dist/leaflet.css\"
        integrity=\"sha256-p4NxAoJBhIIN+hmNHrzRCf9tD/miZyoHS5obTRR9BMY=\"
        crossorigin=\"\"
    />
    <style>
        html, body {{ height: 100%; margin: 0; }}
        #map {{ width: 100%; height: 100%; }}
        .popup-title {{ font-weight: 700; margin-bottom: 4px; }}
        .stop-label {{
            background: transparent;
            border: 0;
            box-shadow: none;
            color: #ffffff;
            font-size: 10px;
            font-weight: 700;
            margin: 0;
            padding: 0;
            text-shadow: 0 0 2px rgba(0, 0, 0, 0.9);
        }}
    </style>
</head>
<body>
    <div id=\"map\"></div>
    <script
        src=\"https://unpkg.com/leaflet@1.9.4/dist/leaflet.js\"
        integrity=\"sha256-20nQCchB9co0qIjJZRGuk2/Z9VM+kNiyxNV1lvTlZBo=\"
        crossorigin=\"\"
    ></script>
    <script>
        const points = {points_json};
        const map = L.map('map').setView([{center_lat}, {center_lon}], 11);

        L.tileLayer('https://tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{
            maxZoom: 19,
            attribution: '&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> contributors'
        }}).addTo(map);

        const latLngs = points.map(p => [p.lat, p.lon]);
        const routeLine = L.polyline(latLngs, {{ color: '#1565c0', weight: 4 }}).addTo(map);

        points.forEach((p, i) => {{
            const isFirst = i === 0;
            const isLast = i === points.length - 1;
            const color = isFirst ? '#2e7d32' : (isLast ? '#c62828' : '#1565c0');
            const marker = L.circleMarker([p.lat, p.lon], {{
                radius: 6,
                color,
                fillColor: color,
                fillOpacity: 0.9
            }}).addTo(map);

            const popup = `<div class=\"popup-title\">Stop ${{p.stop}}</div><div>${{p.address || 'Address unavailable'}}</div>`;
            marker.bindPopup(popup);
            marker.bindTooltip(String(p.stop), {{
                permanent: true,
                direction: 'center',
                className: 'stop-label',
                offset: [0, 0],
                opacity: 1,
            }});
        }});

        if (latLngs.length > 1) {{
            map.fitBounds(routeLine.getBounds(), {{ padding: [30, 30] }});
        }}
    </script>
</body>
</html>
"""

        with open(path, "w", encoding="utf-8") as f:
                f.write(html_doc)

        print(f"  Map saved to: {path}")


# ---------------------------------------------------------------------------
# 5. Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Optimize a delivery route from a CSV or XLSX of addresses."
    )
    parser.add_argument(
        "--input", "-i", default="../RuralMapping.xlsx",
        help="Input CSV/XLSX file with columns: street, city, zip (state optional; default: ../RuralMapping.xlsx)"
    )
    parser.add_argument(
        "--output", "-o", default="route.csv",
        help="Output CSV file with optimized stop order (default: route.csv)"
    )
    parser.add_argument(
        "--map-output", default="route_map.html",
        help="Output Leaflet map HTML file (default: route_map.html)"
    )
    parser.add_argument(
        "--start-index", type=int, default=0,
        help="Row index (0-based) of the starting stop (default: 0)"
    )
    parser.add_argument(
        "--no-two-opt", action="store_true",
        help="Skip 2-opt improvement (faster but lower quality)"
    )
    parser.add_argument(
        "--nominatim", action="store_true",
        help="Use Nominatim/OSM geocoder instead of US Census (for non-US addresses)"
    )
    args = parser.parse_args()

    print("\n=== Route Optimizer ===\n")

    # --- Load ---
    print(f"[1/4] Loading addresses from '{args.input}'...")
    addresses, fieldnames = load_addresses(args.input)
    print(f"  Loaded {len(addresses)} addresses.")

    # --- Geocode ---
    print(f"\n[2/4] Geocoding...")
    manual_coords = extract_manual_coords(addresses)
    needs_geocoding = [i for i, c in enumerate(manual_coords) if c is None]
    manual_count = len(addresses) - len(needs_geocoding)

    if manual_count:
        print(f"  {manual_count} address(es) have manual coordinates and will skip geocoding.")

    if needs_geocoding:
        geocode_subset = [addresses[i] for i in needs_geocoding]
        if args.nominatim:
            geocoded = geocode_nominatim(geocode_subset)
        else:
            geocoded = geocode_census(geocode_subset)
        coords: list[Optional[tuple[float, float]]] = list(manual_coords)
        for subset_idx, orig_idx in enumerate(needs_geocoding):
            coords[orig_idx] = geocoded[subset_idx]
    else:
        print(f"  All addresses have manual coordinates — skipping geocoder.")
        coords = list(manual_coords)

    # Filter out addresses that failed to geocode
    valid_indices = [i for i, c in enumerate(coords) if c is not None]
    failed_count = len(addresses) - len(valid_indices)

    if failed_count > 0:
        print(f"\n  WARNING: {failed_count} addresses could not be geocoded and will be "
              f"appended to the end of the route unchanged.")

    valid_coords = [coords[i] for i in valid_indices]

    if not valid_coords:
        print("ERROR: No addresses were successfully geocoded. Exiting.")
        sys.exit(1)

    # Remap start index to valid-only list
    if args.start_index in valid_indices:
        start = valid_indices.index(args.start_index)
    else:
        print(f"  WARNING: Start index {args.start_index} failed geocoding. Using index 0.")
        start = 0

    # --- Distance Matrix ---
    print(f"\n[3/4] Building distance matrix...")
    dist_matrix = build_distance_matrix(valid_coords)

    # --- Optimize ---
    print(f"\n[4/4] Optimizing route...")
    print(f"  Running nearest-neighbor heuristic (start index: {start})...")
    route = nearest_neighbor(dist_matrix, start=start)
    nn_distance = route_total_distance(route, dist_matrix)
    print(f"  Nearest-neighbor route distance: {nn_distance:.1f} miles")

    if not args.no_two_opt:
        print(f"  Running 2-opt improvement pass...")
        route = two_opt(route, dist_matrix)
        opt_distance = route_total_distance(route, dist_matrix)
        improvement = (nn_distance - opt_distance) / nn_distance * 100
        print(f"  2-opt route distance: {opt_distance:.1f} miles "
              f"({improvement:.1f}% improvement)")
    else:
        print("  Skipping 2-opt (--no-two-opt flag set).")

    # Remap route back to original indices
    final_route = [valid_indices[i] for i in route]

    # Append failed addresses at the end
    failed_indices = [i for i, c in enumerate(coords) if c is None]
    final_route.extend(failed_indices)

    # --- Save ---
    print(f"\nSaving results...")
    save_route(args.output, addresses, final_route, coords, fieldnames)
    save_leaflet_map(args.map_output, addresses, final_route, coords)

    total_distance = route_total_distance(
        [valid_indices.index(i) for i in final_route if i in valid_indices],
        dist_matrix
    )
    print(f"\n=== Done ===")
    print(f"  Stops optimized : {len(valid_indices)}")
    print(f"  Stops appended  : {failed_count} (geocoding failed)")
    print(f"  Total distance  : {total_distance:.1f} miles (straight-line)")
    print(f"  Output file     : {args.output}")
    print(f"  Map file        : {args.map_output}\n")


if __name__ == "__main__":
    main()